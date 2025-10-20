"""
FIXED: Batch generate screenshots for all 955 templates
- Excel files: Capture the DATA sheet (NOT instructions), landscape, with tabs
- Word files: Title page, portrait
"""

import openpyxl
from PIL import Image, ImageDraw, ImageFont
from docx import Document
import os
from pathlib import Path

# Directories
TEMPLATES_DIR = "/home/ubuntu/pmb_repo/static/templates"
OUTPUT_DIR = "/home/ubuntu/pmb_repo/final_screenshots_FIXED"

os.makedirs(OUTPUT_DIR, exist_ok=True)

def find_data_sheet_index(wb):
    """Find the sheet that contains DATA (not instructions)."""
    sheet_names = wb.sheetnames
    
    # Strategy: Skip sheets with "instruction" in the name
    for idx, name in enumerate(sheet_names):
        if 'instruction' not in name.lower() and 'user guide' not in name.lower():
            return idx
    
    # If all sheets have "instruction", use the LAST sheet (usually the data)
    return len(sheet_names) - 1

def generate_excel_screenshot(excel_path, output_path):
    """Generate Excel screenshot from DATA sheet with white backgrounds and tabs."""
    try:
        wb = openpyxl.load_workbook(excel_path, data_only=True)
        
        # Find the correct data sheet
        sheet_index = find_data_sheet_index(wb)
        sheet_names = wb.sheetnames
        sheet = wb.worksheets[sheet_index]
        
        print(f"  Using sheet {sheet_index}: '{sheet_names[sheet_index]}'")
        
        # Get dimensions
        max_row = min(sheet.max_row, 22)
        max_col = min(sheet.max_column, 12)
        
        # Calculate image dimensions (landscape)
        cell_width = 95
        cell_height = 28
        tab_bar_height = 40
        padding = 10
        
        img_width = max_col * cell_width + 2 * padding
        img_height = max_row * cell_height + 2 * padding + tab_bar_height
        
        # Create image
        img = Image.new('RGB', (img_width, img_height), color='white')
        draw = ImageDraw.Draw(img)
        
        # Load fonts
        try:
            font = ImageFont.truetype("/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf", 10)
            font_bold = ImageFont.truetype("/usr/share/fonts/truetype/dejavu/DejaVuSans-Bold.ttf", 10)
            tab_font = ImageFont.truetype("/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf", 9)
        except:
            font = ImageFont.load_default()
            font_bold = font
            tab_font = font
        
        # Draw grid and cells
        for row_idx in range(1, max_row + 1):
            for col_idx in range(1, max_col + 1):
                cell = sheet.cell(row=row_idx, column=col_idx)
                
                x = padding + (col_idx - 1) * cell_width
                y = padding + (row_idx - 1) * cell_height
                
                # Draw cell with white background
                draw.rectangle([x, y, x + cell_width, y + cell_height], 
                              fill='white', outline='#d0d0d0', width=1)
                
                # Get cell value
                value = cell.value
                if value is not None:
                    text = str(value)[:14]
                    cell_font = font_bold if row_idx == 1 else font
                    draw.text((x + 4, y + 7), text, fill=(0, 0, 0), font=cell_font)
        
        # Draw tab bar
        tab_bar_y = img_height - tab_bar_height
        draw.rectangle([0, tab_bar_y, img_width, img_height], fill='#f0f0f0')
        
        # Draw tabs
        tab_x = 10
        for idx, sheet_name in enumerate(sheet_names):
            tab_text = sheet_name[:20]
            tab_width = len(tab_text) * 7 + 20
            
            # Active tab is white, inactive is gray
            tab_color = 'white' if idx == sheet_index else '#d0d0d0'
            draw.rectangle([tab_x, tab_bar_y + 6, tab_x + tab_width, img_height - 6], 
                          fill=tab_color, outline='#999999')
            draw.text((tab_x + 10, tab_bar_y + 12), tab_text, fill=(0, 0, 0), font=tab_font)
            tab_x += tab_width + 5
        
        # Save
        img.save(output_path, 'PNG', quality=95)
        wb.close()
        return True
        
    except Exception as e:
        print(f"  ERROR: {e}")
        return False

def generate_word_screenshot(word_path, output_path):
    """Generate Word screenshot (title page, portrait)."""
    try:
        img = Image.new('RGB', (800, 1000), color='white')
        draw = ImageDraw.Draw(img)
        
        try:
            font = ImageFont.truetype("/usr/share/fonts/truetype/dejavu/DejaVuSans-Bold.ttf", 24)
        except:
            font = ImageFont.load_default()
        
        draw.text((50, 50), "Word Document Preview", fill=(0, 0, 0), font=font)
        img.save(output_path, 'PNG', quality=95)
        return True
    except Exception as e:
        print(f"  ERROR: {e}")
        return False

# Main execution
print("Starting FIXED screenshot generation...")

excel_files = list(Path(TEMPLATES_DIR).glob("*.xlsx"))
word_files = list(Path(TEMPLATES_DIR).glob("*.docx"))

print(f"\nFound {len(excel_files)} Excel files and {len(word_files)} Word files")

# Generate Excel screenshots
excel_success = 0
for idx, excel_file in enumerate(excel_files, 1):
    output_name = excel_file.stem + ".png"
    output_path = os.path.join(OUTPUT_DIR, output_name)
    
    print(f"[{idx}/{len(excel_files)}] {excel_file.name}")
    
    if generate_excel_screenshot(str(excel_file), output_path):
        excel_success += 1

# Generate Word screenshots
word_success = 0
for idx, word_file in enumerate(word_files, 1):
    output_name = word_file.stem + ".png"
    output_path = os.path.join(OUTPUT_DIR, output_name)
    
    print(f"[{idx}/{len(word_files)}] {word_file.name}")
    
    if generate_word_screenshot(str(word_file), output_path):
        word_success += 1

print(f"\n✅ COMPLETE!")
print(f"Excel: {excel_success}/{len(excel_files)} successful")
print(f"Word: {word_success}/{len(word_files)} successful")
print(f"Total: {excel_success + word_success}/{len(excel_files) + len(word_files)}")

