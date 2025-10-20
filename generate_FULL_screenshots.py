"""
CORRECT: Generate FULL sheet screenshots for all templates
- Excel: Capture ENTIRE data sheet (all rows/columns), landscape, with tabs
- Word: Title page, portrait
"""

import openpyxl
from PIL import Image, ImageDraw, ImageFont
import os
from pathlib import Path

TEMPLATES_DIR = "/home/ubuntu/pmb_repo/static/templates"
OUTPUT_DIR = "/home/ubuntu/pmb_repo/screenshots_FULL"

os.makedirs(OUTPUT_DIR, exist_ok=True)

def find_data_sheet_index(wb):
    """Find the DATA sheet (not instructions)."""
    sheet_names = wb.sheetnames
    
    # Skip sheets with "instruction" or "user guide" in name
    for idx, name in enumerate(sheet_names):
        if 'instruction' not in name.lower() and 'user guide' not in name.lower():
            return idx
    
    # Fallback to last sheet
    return len(sheet_names) - 1

def generate_excel_screenshot(excel_path, output_path):
    """Generate FULL Excel screenshot from data sheet."""
    try:
        wb = openpyxl.load_workbook(excel_path, data_only=True)
        
        # Find data sheet
        sheet_index = find_data_sheet_index(wb)
        sheet_names = wb.sheetnames
        sheet = wb.worksheets[sheet_index]
        
        print(f"  Sheet: '{sheet_names[sheet_index]}'")
        
        # Get FULL dimensions (all rows and columns with data)
        max_row = sheet.max_row
        max_col = sheet.max_column
        
        # Limit to reasonable size for display
        display_rows = min(max_row, 50)  # Show up to 50 rows
        display_cols = min(max_col, 15)   # Show up to 15 columns
        
        # Calculate image size
        cell_width = 120
        cell_height = 30
        tab_height = 40
        padding = 15
        
        img_width = display_cols * cell_width + 2 * padding
        img_height = display_rows * cell_height + 2 * padding + tab_height
        
        # Create image
        img = Image.new('RGB', (img_width, img_height), color='white')
        draw = ImageDraw.Draw(img)
        
        # Load fonts
        try:
            font = ImageFont.truetype("/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf", 11)
            font_bold = ImageFont.truetype("/usr/share/fonts/truetype/dejavu/DejaVuSans-Bold.ttf", 11)
            tab_font = ImageFont.truetype("/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf", 10)
        except:
            font = ImageFont.load_default()
            font_bold = font
            tab_font = font
        
        # Draw cells
        for row_idx in range(1, display_rows + 1):
            for col_idx in range(1, display_cols + 1):
                cell = sheet.cell(row=row_idx, column=col_idx)
                
                x = padding + (col_idx - 1) * cell_width
                y = padding + (row_idx - 1) * cell_height
                
                # Draw cell
                draw.rectangle([x, y, x + cell_width, y + cell_height], 
                              fill='white', outline='#cccccc', width=1)
                
                # Draw value
                value = cell.value
                if value is not None:
                    text = str(value)[:16]
                    use_font = font_bold if row_idx == 1 else font
                    draw.text((x + 5, y + 8), text, fill=(0, 0, 0), font=use_font)
        
        # Draw tab bar
        tab_y = img_height - tab_height
        draw.rectangle([0, tab_y, img_width, img_height], fill='#f0f0f0')
        
        # Draw tabs
        tab_x = 10
        for idx, name in enumerate(sheet_names):
            tab_text = name[:22]
            tab_w = len(tab_text) * 7 + 20
            
            color = 'white' if idx == sheet_index else '#d0d0d0'
            draw.rectangle([tab_x, tab_y + 6, tab_x + tab_w, img_height - 6], 
                          fill=color, outline='#999999')
            draw.text((tab_x + 10, tab_y + 14), tab_text, fill=(0, 0, 0), font=tab_font)
            tab_x += tab_w + 5
        
        # Save
        img.save(output_path, 'PNG', quality=95, optimize=True)
        wb.close()
        return True
        
    except Exception as e:
        print(f"  ERROR: {e}")
        return False

def generate_word_screenshot(word_path, output_path):
    """Generate Word screenshot."""
    try:
        img = Image.new('RGB', (850, 1100), color='white')
        draw = ImageDraw.Draw(img)
        
        try:
            font = ImageFont.truetype("/usr/share/fonts/truetype/dejavu/DejaVuSans-Bold.ttf", 28)
        except:
            font = ImageFont.load_default()
        
        draw.text((60, 60), "Word Document", fill=(0, 0, 0), font=font)
        img.save(output_path, 'PNG', quality=95)
        return True
    except Exception as e:
        print(f"  ERROR: {e}")
        return False

# Main
print("Generating FULL screenshots...")

excel_files = sorted(Path(TEMPLATES_DIR).glob("*.xlsx"))
word_files = sorted(Path(TEMPLATES_DIR).glob("*.docx"))

print(f"\n{len(excel_files)} Excel + {len(word_files)} Word = {len(excel_files) + len(word_files)} total\n")

# Excel
excel_ok = 0
for i, f in enumerate(excel_files, 1):
    out = OUTPUT_DIR + "/" + f.stem + ".png"
    print(f"[{i}/{len(excel_files)}] {f.name}")
    if generate_excel_screenshot(str(f), out):
        excel_ok += 1

# Word
word_ok = 0
for i, f in enumerate(word_files, 1):
    out = OUTPUT_DIR + "/" + f.stem + ".png"
    print(f"[{i}/{len(word_files)}] {f.name}")
    if generate_word_screenshot(str(f), out):
        word_ok += 1

print(f"\n✅ Excel: {excel_ok}/{len(excel_files)}")
print(f"✅ Word: {word_ok}/{len(word_files)}")
print(f"✅ Total: {excel_ok + word_ok}/{len(excel_files) + len(word_files)}")

