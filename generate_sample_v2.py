#!/usr/bin/env python3
"""
Generate high-quality Excel screenshot from sheet 2 (dashboard) in landscape orientation.
Uses better rendering to show actual cell content and formatting.
"""

import openpyxl
from openpyxl.styles import Color
from PIL import Image, ImageDraw, ImageFont
import os
import sys

def rgb_from_excel_color(color_obj):
    """Convert Excel color object to RGB tuple."""
    if not color_obj:
        return (255, 255, 255)  # White default
    
    try:
        if hasattr(color_obj, 'rgb') and color_obj.rgb:
            rgb_str = color_obj.rgb
            if len(rgb_str) == 8:  # ARGB format
                rgb_str = rgb_str[2:]  # Skip alpha
            r = int(rgb_str[0:2], 16)
            g = int(rgb_str[2:4], 16)
            b = int(rgb_str[4:6], 16)
            return (r, g, b)
    except:
        pass
    
    return (255, 255, 255)  # White default


def generate_excel_screenshot_v2(excel_path, output_path, sheet_index=1):
    """
    Generate a high-quality landscape screenshot of an Excel sheet.
    
    Args:
        excel_path: Path to the Excel file
        output_path: Path to save the screenshot
        sheet_index: Index of sheet to capture (0-based, default 1 for second sheet)
    """
    print(f"Loading Excel file: {excel_path}")
    
    # Load the workbook with data_only=True to get calculated values
    wb = openpyxl.load_workbook(excel_path, data_only=True)
    
    # Get the sheet names
    sheet_names = wb.sheetnames
    print(f"Available sheets: {sheet_names}")
    
    if sheet_index >= len(sheet_names):
        print(f"Warning: Sheet index {sheet_index} not found, using first sheet")
        sheet_index = 0
    
    sheet = wb.worksheets[sheet_index]
    print(f"Using sheet: {sheet.title} (index {sheet_index})")
    
    # Get the used range
    max_row = sheet.max_row
    max_col = sheet.max_column
    print(f"Sheet dimensions: {max_row} rows x {max_col} columns")
    
    # Calculate image dimensions (landscape)
    cell_width = 140  # pixels per cell (wider for better readability)
    cell_height = 35  # pixels per cell
    
    # Limit to reasonable size (first 20 rows, all columns)
    display_rows = min(max_row, 20)
    display_cols = max_col
    
    img_width = display_cols * cell_width
    img_height = display_rows * cell_height + 60  # Extra space for header
    
    print(f"Creating image: {img_width}x{img_height} pixels (landscape)")
    
    # Create image with white background
    img = Image.new('RGB', (img_width, img_height), color='white')
    draw = ImageDraw.Draw(img)
    
    # Try to use a nice font
    try:
        font = ImageFont.truetype("/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf", 11)
        header_font = ImageFont.truetype("/usr/share/fonts/truetype/dejavu/DejaVuSans-Bold.ttf", 13)
    except:
        font = ImageFont.load_default()
        header_font = ImageFont.load_default()
    
    # Draw sheet name header
    header_text = f"Sheet: {sheet.title}"
    draw.rectangle([0, 0, img_width, 45], fill='#34495e')
    draw.text((15, 12), header_text, fill='white', font=header_font)
    
    y_offset = 50
    
    # Draw grid and cell contents
    for row_idx in range(1, display_rows + 1):
        for col_idx in range(1, display_cols + 1):
            cell = sheet.cell(row=row_idx, column=col_idx)
            
            # Calculate cell position
            x = (col_idx - 1) * cell_width
            y = y_offset + (row_idx - 1) * cell_height
            
            # Get background color
            bg_color = (255, 255, 255)  # White default
            if cell.fill and cell.fill.start_color:
                bg_color = rgb_from_excel_color(cell.fill.start_color)
            
            # Fill cell background
            draw.rectangle([x, y, x + cell_width, y + cell_height], 
                          fill=bg_color, outline='#d0d0d0', width=1)
            
            # Get cell value
            value = cell.value
            if value is not None:
                # Convert to string
                text = str(value)
                
                # Truncate if too long
                max_chars = 18
                if len(text) > max_chars:
                    text = text[:max_chars-3] + "..."
                
                # Get text color (default to black, or white if background is dark)
                text_color = (0, 0, 0)  # Black default
                
                # Check if we have a font color
                if cell.font and cell.font.color:
                    text_color = rgb_from_excel_color(cell.font.color)
                else:
                    # Auto-detect: use white text on dark backgrounds
                    brightness = (bg_color[0] * 299 + bg_color[1] * 587 + bg_color[2] * 114) / 1000
                    if brightness < 128:
                        text_color = (255, 255, 255)
                
                # Draw text with some padding
                draw.text((x + 8, y + 10), text, fill=text_color, font=font)
    
    # Save the image
    img.save(output_path, 'PNG', quality=95, optimize=True)
    print(f"Screenshot saved to: {output_path}")
    print(f"Image size: {img.width}x{img.height} pixels")
    
    return output_path


if __name__ == "__main__":
    # Test with KPI Dashboard file
    test_file = "/home/ubuntu/pmb_repo/static/templates/AI_ML_KPI_Dashboard_with_Instructions_2025_PMI.xlsx"
    output_file = "/home/ubuntu/pmb_repo/sample_screenshot_v2.png"
    
    if len(sys.argv) > 1:
        test_file = sys.argv[1]
    if len(sys.argv) > 2:
        output_file = sys.argv[2]
    
    if not os.path.exists(test_file):
        print(f"Error: File not found: {test_file}")
        sys.exit(1)
    
    generate_excel_screenshot_v2(test_file, output_file, sheet_index=1)
    print("\n✓ Sample screenshot generated successfully!")
    print(f"  View at: {output_file}")

