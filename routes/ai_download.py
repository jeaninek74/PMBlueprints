"""
AI Template Download Route
Generates and downloads AI-created templates
"""

from flask import Blueprint, request, jsonify, send_file
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
import io
import logging

logger = logging.getLogger(__name__)

ai_download_bp = Blueprint('ai_download', __name__)

@ai_download_bp.route('/api/ai/download-generated', methods=['POST'])
def download_generated_template():
    """
    Generate and download an AI-created Excel template
    """
    try:
        data = request.get_json()
        template_name = data.get('template_name', 'AI_Generated_Template')
        description = data.get('description', '')
        sections = data.get('sections', [])
        
        # Create Excel workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Template"
        
        # Header styling
        header_fill = PatternFill(start_color="1E3A8A", end_color="1E3A8A", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True, size=12)
        
        # Title
        ws['A1'] = template_name
        ws['A1'].font = Font(bold=True, size=16, color="1E3A8A")
        ws.merge_cells('A1:D1')
        
        # Description
        ws['A2'] = description
        ws.merge_cells('A2:D2')
        ws['A2'].alignment = Alignment(wrap_text=True)
        
        # Add sections
        row = 4
        ws[f'A{row}'] = "Section"
        ws[f'B{row}'] = "Description"
        ws[f'C{row}'] = "Status"
        ws[f'D{row}'] = "Notes"
        
        # Style header row
        for col in ['A', 'B', 'C', 'D']:
            cell = ws[f'{col}{row}']
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')
        
        # Add section rows
        row += 1
        for section in sections:
            ws[f'A{row}'] = section
            ws[f'B{row}'] = f"Details for {section}"
            ws[f'C{row}'] = "Not Started"
            ws[f'D{row}'] = ""
            row += 1
        
        # Adjust column widths
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 40
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 30
        
        # Save to BytesIO
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        # Generate filename
        filename = f"{template_name.replace(' ', '_')}.xlsx"
        
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )
        
    except Exception as e:
        logger.error(f"Error generating template: {e}")
        return jsonify({'error': str(e)}), 500

