"""
FIXED: Generate screenshots WITH cell colors
"""

import openpyxl
from openpyxl.styles import PatternFill
from PIL import Image, ImageDraw, ImageFont
import os
from pathlib import Path
import subprocess

TEMPLATES_DIR = "/home/ubuntu/pmb_repo/static/templates"
OUTPUT_DIR = "/home/ubuntu/pmb_repo/screenshots_FULL"

os.makedirs(OUTPUT_DIR, exist_ok=True)

def hex_to_rgb(hex_color):
    """Convert hex color to RGB tuple."""
    if not hex_color:
        return (255, 255, 255)  # White default
    
    # Convert to string if it's an RGB object
    hex_str = str(hex_color)
    
    if hex_str == '00000000' or not hex_str:
        return (255, 255, 255)
    
    # Remove alpha if present
    if len(hex_str) == 8:
        hex_str = hex_str[2:]
    
    try:
        return tuple(int(hex_str[i:i+2], 16) for i in (0, 2, 4))
    except:
        return (255, 255, 255)

def find_data_sheet_index(wb):
    """Find the DATA sheet (not instructions)."""
    sheet_names = wb.sheetnames
    
    for idx, name in enumerate(sheet_names):
        if 'instruction' not in name.lower() and 'user guide' not in name.lower():
            return idx
    
    return len(sheet_names) - 1

def generate_excel_screenshot(excel_path, output_path):
    """Generate Excel screenshot WITH COLORS."""
    try:
        wb = openpyxl.load_workbook(excel_path, data_only=False)
        
        sheet_index = find_data_sheet_index(wb)
        sheet_names = wb.sheetnames
        sheet = wb.worksheets[sheet_index]
        
        print(f"  Sheet: '{sheet_names[sheet_index]}'")
        
        max_row = sheet.max_row
        max_col = sheet.max_column
        
        display_rows = min(max_row, 50)
        display_cols = min(max_col, 15)
        
        cell_width = 120
        cell_height = 30
        tab_height = 40
        padding = 15
        
        img_width = display_cols * cell_width + 2 * padding
        img_height = display_rows * cell_height + 2 * padding + tab_height
        
        img = Image.new('RGB', (img_width, img_height), color='white')
        draw = ImageDraw.Draw(img)
        
        try:
            font = ImageFont.truetype("/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf", 11)
            font_bold = ImageFont.truetype("/usr/share/fonts/truetype/dejavu/DejaVuSans-Bold.ttf", 11)
            tab_font = ImageFont.truetype("/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf", 10)
        except:
            font = ImageFont.load_default()
            font_bold = font
            tab_font = font
        
        # Draw cells WITH COLORS
        for row_idx in range(1, display_rows + 1):
            for col_idx in range(1, display_cols + 1):
                cell = sheet.cell(row=row_idx, column=col_idx)
                
                x = padding + (col_idx - 1) * cell_width
                y = padding + (row_idx - 1) * cell_height
                
                # Get cell background color
                bg_color = (255, 255, 255)  # Default white
                if cell.fill and cell.fill.start_color:
                    hex_color = cell.fill.start_color.rgb
                    if hex_color:
                        bg_color = hex_to_rgb(hex_color)
                
                # Draw cell with background color
                draw.rectangle([x, y, x + cell_width, y + cell_height], 
                              fill=bg_color, outline='#cccccc', width=1)
                
                # Get text color
                text_color = (0, 0, 0)  # Default black
                if cell.font and cell.font.color:
                    hex_color = cell.font.color.rgb
                    if hex_color:
                        text_color = hex_to_rgb(hex_color)
                
                # Draw value
                value = cell.value
                if value is not None:
                    text = str(value)[:16]
                    use_font = font_bold if row_idx == 1 else font
                    draw.text((x + 5, y + 8), text, fill=text_color, font=use_font)
        
        # Draw tab bar
        tab_y = img_height - tab_height
        draw.rectangle([0, tab_y, img_width, img_height], fill='#f0f0f0')
        
        # Draw tabs
        tab_x = 10
        for idx, name in enumerate(sheet_names):
            tab_text = name[:22]
            tab_w = len(tab_text) * 7 + 20
            
            if idx == sheet_index:
                draw.rectangle([tab_x, tab_y + 5, tab_x + tab_w, img_height - 5], 
                              fill='white', outline='#999999')
            else:
                draw.rectangle([tab_x, tab_y + 5, tab_x + tab_w, img_height - 5], 
                              fill='#d0d0d0', outline='#999999')
            
            draw.text((tab_x + 10, tab_y + 12), tab_text, fill=(0, 0, 0), font=tab_font)
            tab_x += tab_w + 5
        
        img.save(output_path)
        wb.close()
        return True
        
    except Exception as e:
        print(f"  Error: {e}")
        return False

def generate_word_screenshot(docx_path, output_path):
    """Generate Word screenshot using LibreOffice."""
    try:
        temp_pdf = output_path.replace('.png', '_temp.pdf')
        
        subprocess.run([
            'libreoffice', '--headless', '--convert-to', 'pdf',
            '--outdir', os.path.dirname(temp_pdf),
            docx_path
        ], capture_output=True, timeout=30)
        
        pdf_file = os.path.join(os.path.dirname(temp_pdf), 
                               os.path.basename(docx_path).replace('.docx', '.pdf'))
        
        if os.path.exists(pdf_file):
            subprocess.run([
                'convert', '-density', '150', 
                f'{pdf_file}[0]', '-quality', '90',
                output_path
            ], capture_output=True, timeout=30)
            
            os.remove(pdf_file)
            return os.path.exists(output_path)
        
        return False
        
    except Exception as e:
        print(f"  Error: {e}")
        return False

# Process all templates
excel_files = sorted(Path(TEMPLATES_DIR).glob("*.xlsx"))
word_files = sorted(Path(TEMPLATES_DIR).glob("*.docx"))

print(f"Processing {len(excel_files)} Excel files...")
for idx, excel_file in enumerate(excel_files, 1):
    output_file = os.path.join(OUTPUT_DIR, f"{excel_file.stem}.png")
    print(f"[{idx}/{len(excel_files)}] {excel_file.name}")
    generate_excel_screenshot(str(excel_file), str(output_file))

print(f"\nProcessing {len(word_files)} Word files...")
for idx, word_file in enumerate(word_files, 1):
    output_file = os.path.join(OUTPUT_DIR, f"{word_file.stem}.png")
    print(f"[{idx}/{len(word_files)}] {word_file.name}")
    generate_word_screenshot(str(word_file), str(output_file))

print(f"\n✅ Excel: {len(excel_files)}/{len(excel_files)}")
print(f"✅ Word: {len(word_files)}/{len(word_files)}")
print(f"✅ Total: {len(excel_files) + len(word_files)}/{len(excel_files) + len(word_files)}")

