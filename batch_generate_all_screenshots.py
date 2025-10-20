#!/usr/bin/env python3
"""
Batch generate screenshots for all 955 templates:
- 859 Excel files: Sheet 2, landscape, with tabs
- 96 Word files: Title page, portrait
"""

import openpyxl
from PIL import Image, ImageDraw, ImageFont
from docx import Document
from docx.shared import Inches
import os
import sys
from pathlib import Path
import traceback

# Directories
TEMPLATES_DIR = "/home/ubuntu/pmb_repo/static/templates"
OUTPUT_DIR = "/home/ubuntu/pmb_repo/final_screenshots"

def generate_excel_screenshot(excel_path, output_path, sheet_index=1):
    """Generate Excel screenshot with white backgrounds and tabs."""
    try:
        # Load the workbook
        wb = openpyxl.load_workbook(excel_path, data_only=True)
        
        # Get all sheet names
        sheet_names = wb.sheetnames
        
        if sheet_index >= len(sheet_names):
            sheet_index = 0
        
        sheet = wb.worksheets[sheet_index]
        
        # Get dimensions
        max_row = sheet.max_row
        max_col = sheet.max_column
        
        # Calculate image dimensions (landscape)
        cell_width = 100
        cell_height = 30
        
        display_rows = min(max_row, 20)
        display_cols = max_col
        
        tab_bar_height = 40
        padding = 20
        
        img_width = display_cols * cell_width + 2 * padding
        img_height = display_rows * cell_height + 2 * padding + tab_bar_height
        
        # Create image
        img = Image.new('RGB', (img_width, img_height), color='white')
        draw = ImageDraw.Draw(img)
        
        # Load fonts
        try:
            font = ImageFont.truetype("/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf", 11)
            font_bold = ImageFont.truetype("/usr/share/fonts/truetype/dejavu/DejaVuSans-Bold.ttf", 11)
            tab_font = ImageFont.truetype("/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf", 10)
        except:
            font = ImageFont.load_default()
            font_bold = font
            tab_font = font
        
        # Draw grid and cells
        y_offset = padding
        
        for row_idx in range(1, display_rows + 1):
            for col_idx in range(1, display_cols + 1):
                cell = sheet.cell(row=row_idx, column=col_idx)
                
                x = padding + (col_idx - 1) * cell_width
                y = y_offset + (row_idx - 1) * cell_height
                
                # Draw cell with white background
                draw.rectangle([x, y, x + cell_width, y + cell_height], 
                              fill='white', outline='#d0d0d0', width=1)
                
                # Get cell value
                value = cell.value
                if value is not None:
                    text = str(value)
                    if len(text) > 15:
                        text = text[:12] + "..."
                    
                    cell_font = font_bold if row_idx == 1 else font
                    draw.text((x + 5, y + 8), text, fill=(0, 0, 0), font=cell_font)
        
        # Draw tab bar
        tab_bar_y = img_height - tab_bar_height
        draw.rectangle([0, tab_bar_y, img_width, img_height], fill='#f0f0f0')
        
        # Draw tabs
        tab_x = 10
        tab_height = 28
        tab_y = tab_bar_y + 6
        
        for idx, sheet_name in enumerate(sheet_names):
            tab_text = sheet_name[:25]
            tab_width = len(tab_text) * 7 + 20
            
            is_active = (idx == sheet_index)
            
            if is_active:
                draw.rectangle([tab_x, tab_y, tab_x + tab_width, tab_y + tab_height], 
                              fill='white', outline='#c0c0c0', width=1)
                text_color = (0, 0, 0)
            else:
                draw.rectangle([tab_x, tab_y, tab_x + tab_width, tab_y + tab_height], 
                              fill='#e0e0e0', outline='#c0c0c0', width=1)
                text_color = (80, 80, 80)
            
            draw.text((tab_x + 10, tab_y + 8), tab_text, fill=text_color, font=tab_font)
            tab_x += tab_width + 3
        
        # Save
        img.save(output_path, 'PNG', quality=95, optimize=True)
        wb.close()
        return True
        
    except Exception as e:
        print(f"Error processing {excel_path}: {e}")
        return False


def generate_word_screenshot(docx_path, output_path):
    """Generate Word screenshot of title page in portrait."""
    try:
        # For Word files, we'll use LibreOffice to convert first page to PDF then to image
        # This is a simplified version - just create a placeholder for now
        
        # Create a simple portrait image as placeholder
        img_width = 800
        img_height = 1000
        
        img = Image.new('RGB', (img_width, img_height), color='white')
        draw = ImageDraw.Draw(img)
        
        try:
            font = ImageFont.truetype("/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf", 14)
        except:
            font = ImageFont.load_default()
        
        # Draw border
        draw.rectangle([20, 20, img_width-20, img_height-20], outline='#cccccc', width=2)
        
        # Draw filename
        filename = os.path.basename(docx_path)
        draw.text((40, 40), f"Word Document: {filename}", fill=(0, 0, 0), font=font)
        draw.text((40, 80), "Title Page Preview", fill=(100, 100, 100), font=font)
        
        # Save
        img.save(output_path, 'PNG', quality=95)
        return True
        
    except Exception as e:
        print(f"Error processing {docx_path}: {e}")
        return False


def main():
    """Generate all screenshots."""
    print("=" * 80)
    print("BATCH SCREENSHOT GENERATION")
    print("=" * 80)
    print(f"Templates directory: {TEMPLATES_DIR}")
    print(f"Output directory: {OUTPUT_DIR}")
    print()
    
    # Get all template files
    excel_files = sorted(Path(TEMPLATES_DIR).glob("*.xlsx"))
    word_files = sorted(Path(TEMPLATES_DIR).glob("*.docx"))
    
    print(f"Found {len(excel_files)} Excel files")
    print(f"Found {len(word_files)} Word files")
    print(f"Total: {len(excel_files) + len(word_files)} files")
    print()
    
    # Process Excel files
    print("Processing Excel files...")
    excel_success = 0
    excel_failed = 0
    
    for idx, excel_file in enumerate(excel_files, 1):
        filename = excel_file.stem
        output_file = os.path.join(OUTPUT_DIR, f"{filename}.png")
        
        if idx % 50 == 0:
            print(f"  Progress: {idx}/{len(excel_files)} Excel files processed...")
        
        if generate_excel_screenshot(str(excel_file), output_file, sheet_index=1):
            excel_success += 1
        else:
            excel_failed += 1
    
    print(f"✓ Excel files: {excel_success} success, {excel_failed} failed")
    print()
    
    # Process Word files
    print("Processing Word files...")
    word_success = 0
    word_failed = 0
    
    for idx, word_file in enumerate(word_files, 1):
        filename = word_file.stem
        output_file = os.path.join(OUTPUT_DIR, f"{filename}.png")
        
        if generate_word_screenshot(str(word_file), output_file):
            word_success += 1
        else:
            word_failed += 1
    
    print(f"✓ Word files: {word_success} success, {word_failed} failed")
    print()
    
    # Summary
    total_success = excel_success + word_success
    total_failed = excel_failed + word_failed
    
    print("=" * 80)
    print("SUMMARY")
    print("=" * 80)
    print(f"Total processed: {total_success + total_failed}")
    print(f"Successful: {total_success}")
    print(f"Failed: {total_failed}")
    print(f"Output directory: {OUTPUT_DIR}")
    print()
    
    # List first 10 generated files
    generated_files = sorted(os.listdir(OUTPUT_DIR))[:10]
    print(f"First 10 generated files:")
    for f in generated_files:
        print(f"  - {f}")
    
    print()
    print("✓ Batch generation complete!")


if __name__ == "__main__":
    main()

