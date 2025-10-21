#!/usr/bin/env python3
"""
Automated Screenshot Capture Script for PM Blueprints Templates
Captures screenshots from Excel (Dashboard tab, landscape) and Word (title page, portrait) files
"""

import os
import sys
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.drawing.image import Image as XLImage
from docx import Document
from PIL import Image, ImageDraw, ImageFont
import io

# Template directory
TEMPLATE_DIR = Path("static/templates")
OUTPUT_DIR = Path("static/screenshots_new")

def create_excel_screenshot(excel_path, output_path):
    """
    Create a screenshot of the Excel Dashboard tab (2nd tab)
    Returns True if successful, False otherwise
    """
    try:
        wb = load_workbook(excel_path, data_only=True)
        
        # Get the Dashboard tab (second sheet, index 1)
        if len(wb.sheetnames) < 2:
            print(f"  ⚠️  {excel_path.name} has less than 2 sheets, using first sheet")
            sheet = wb.worksheets[0]
        else:
            sheet = wb.worksheets[1]  # Second tab (Dashboard)
        
        print(f"  📊 Capturing sheet: {sheet.title}")
        
        # Calculate dimensions
        max_row = sheet.max_row
        max_col = sheet.max_column
        
        # Create a landscape image (wider than tall)
        # Approximate: 100 pixels per column, 25 pixels per row
        width = max(1200, min(max_col * 100, 2400))  # Min 1200px, max 2400px
        height = max(800, min(max_row * 25, 1600))   # Min 800px, max 1600px
        
        # Create image with white background
        img = Image.new('RGB', (width, height), color='white')
        draw = ImageDraw.Draw(img)
        
        # Try to load a font
        try:
            font = ImageFont.truetype("/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf", 12)
            title_font = ImageFont.truetype("/usr/share/fonts/truetype/dejavu/DejaVuSans-Bold.ttf", 16)
        except:
            font = ImageFont.load_default()
            title_font = ImageFont.load_default()
        
        # Draw title
        title = f"{sheet.title} - {excel_path.stem}"
        draw.text((20, 20), title, fill='black', font=title_font)
        
        # Draw a simple representation
        y_offset = 60
        x_offset = 20
        
        # Draw column headers (first row)
        for col_idx, col in enumerate(sheet.iter_cols(max_col=min(max_col, 12), max_row=1)):
            cell = col[0]
            if cell.value:
                draw.text((x_offset + col_idx * 100, y_offset), 
                         str(cell.value)[:15], fill='#1f4788', font=font)
        
        # Draw some data rows
        y_offset += 30
        for row_idx, row in enumerate(sheet.iter_rows(min_row=2, max_row=min(max_row, 20), 
                                                       max_col=min(max_col, 12))):
            for col_idx, cell in enumerate(row):
                if cell.value:
                    # Color code based on value type
                    color = 'black'
                    if isinstance(cell.value, (int, float)):
                        color = '#006400'  # Dark green for numbers
                    
                    text = str(cell.value)[:15]
                    draw.text((x_offset + col_idx * 100, y_offset + row_idx * 25), 
                             text, fill=color, font=font)
        
        # Draw tab bar at bottom
        tab_y = height - 40
        draw.rectangle([(0, tab_y), (width, height)], fill='#f0f0f0')
        for idx, sheet_name in enumerate(wb.sheetnames[:5]):
            tab_color = '#4472C4' if idx == 1 else '#808080'
            draw.text((20 + idx * 150, tab_y + 10), sheet_name, fill=tab_color, font=font)
        
        # Save the image
        img.save(output_path, 'PNG', quality=95)
        print(f"  ✅ Saved screenshot: {output_path.name}")
        return True
        
    except Exception as e:
        print(f"  ❌ Error capturing Excel screenshot: {e}")
        return False

def create_word_screenshot(word_path, output_path):
    """
    Create a screenshot of the Word document title page
    Returns True if successful, False otherwise
    """
    try:
        doc = Document(word_path)
        
        # Create a portrait image (taller than wide)
        width = 850  # Portrait width
        height = 1100  # Portrait height
        
        # Create image with white background
        img = Image.new('RGB', (width, height), color='white')
        draw = ImageDraw.Draw(img)
        
        # Try to load fonts
        try:
            title_font = ImageFont.truetype("/usr/share/fonts/truetype/dejavu/DejaVuSans-Bold.ttf", 24)
            subtitle_font = ImageFont.truetype("/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf", 16)
            body_font = ImageFont.truetype("/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf", 12)
        except:
            title_font = ImageFont.load_default()
            subtitle_font = ImageFont.load_default()
            body_font = ImageFont.load_default()
        
        # Extract text from first page (first 10 paragraphs)
        y_offset = 100
        x_offset = 60
        
        for idx, para in enumerate(doc.paragraphs[:15]):
            if para.text.strip():
                # Determine font based on paragraph style
                if idx == 0 or 'title' in para.style.name.lower():
                    font = title_font
                    color = '#1f4788'
                elif 'heading' in para.style.name.lower() or idx < 3:
                    font = subtitle_font
                    color = '#2e5090'
                else:
                    font = body_font
                    color = 'black'
                
                # Wrap text
                text = para.text[:80]
                if len(para.text) > 80:
                    text += "..."
                
                draw.text((x_offset, y_offset), text, fill=color, font=font)
                y_offset += 40 if font == title_font else (30 if font == subtitle_font else 25)
                
                if y_offset > height - 100:
                    break
        
        # Add document info at bottom
        draw.text((x_offset, height - 60), 
                 f"Document: {word_path.stem}", fill='#666666', font=body_font)
        draw.text((x_offset, height - 40), 
                 f"Pages: {len(doc.sections)}", fill='#666666', font=body_font)
        
        # Save the image
        img.save(output_path, 'PNG', quality=95)
        print(f"  ✅ Saved screenshot: {output_path.name}")
        return True
        
    except Exception as e:
        print(f"  ❌ Error capturing Word screenshot: {e}")
        return False

def main():
    """Main function to process all templates"""
    
    # Create output directory
    OUTPUT_DIR.mkdir(exist_ok=True)
    
    print("=" * 70)
    print("PM BLUEPRINTS - AUTOMATED SCREENSHOT CAPTURE")
    print("=" * 70)
    print()
    
    # Find all Excel and Word files
    excel_files = list(TEMPLATE_DIR.glob("*.xlsx")) + list(TEMPLATE_DIR.glob("*.xls"))
    word_files = list(TEMPLATE_DIR.glob("*.docx")) + list(TEMPLATE_DIR.glob("*.doc"))
    
    total_files = len(excel_files) + len(word_files)
    print(f"📁 Found {len(excel_files)} Excel files and {len(word_files)} Word files")
    print(f"📊 Total templates to process: {total_files}")
    print()
    
    success_count = 0
    error_count = 0
    
    # Process Excel files
    print("Processing Excel files...")
    print("-" * 70)
    for idx, excel_file in enumerate(excel_files, 1):
        print(f"[{idx}/{len(excel_files)}] {excel_file.name}")
        output_file = OUTPUT_DIR / f"{excel_file.stem}.png"
        
        if create_excel_screenshot(excel_file, output_file):
            success_count += 1
        else:
            error_count += 1
        print()
    
    # Process Word files
    print("Processing Word files...")
    print("-" * 70)
    for idx, word_file in enumerate(word_files, 1):
        print(f"[{idx}/{len(word_files)}] {word_file.name}")
        output_file = OUTPUT_DIR / f"{word_file.stem}.png"
        
        if create_word_screenshot(word_file, output_file):
            success_count += 1
        else:
            error_count += 1
        print()
    
    # Summary
    print("=" * 70)
    print("SUMMARY")
    print("=" * 70)
    print(f"✅ Successfully captured: {success_count} screenshots")
    print(f"❌ Errors: {error_count}")
    print(f"📁 Output directory: {OUTPUT_DIR.absolute()}")
    print()
    
    return 0 if error_count == 0 else 1

if __name__ == "__main__":
    sys.exit(main())

