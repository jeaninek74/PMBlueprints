#!/usr/bin/env python3
"""
Template Audit Script
Checks all 955 templates to verify file content matches template name
"""

import os
import sys
import zipfile
import openpyxl
from pathlib import Path

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

from app import app, db
from models import Template

def extract_docx_title(file_path):
    """Extract title from DOCX file"""
    try:
        with zipfile.ZipFile(file_path, 'r') as zip_ref:
            xml_content = zip_ref.read('word/document.xml').decode('utf-8', errors='ignore')
            
            # Look for common title patterns
            import re
            # Find text between <w:t> tags in first 5000 chars
            matches = re.findall(r'<w:t[^>]*>([^<]+)</w:t>', xml_content[:5000])
            
            if matches:
                # Join first few matches to get title
                title = ' '.join(matches[:10])
                return title[:200]  # First 200 chars
            return "NO TITLE FOUND"
    except Exception as e:
        return f"ERROR: {str(e)}"

def extract_xlsx_title(file_path):
    """Extract title from XLSX file"""
    try:
        wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
        ws = wb[wb.sheetnames[0]]
        
        # Get first few cells
        title_parts = []
        for row in ws.iter_rows(min_row=1, max_row=5, values_only=True):
            for cell in row:
                if cell and isinstance(cell, str) and len(cell.strip()) > 0:
                    title_parts.append(cell.strip())
                    if len(title_parts) >= 3:
                        break
            if len(title_parts) >= 3:
                break
        
        return ' | '.join(title_parts[:3]) if title_parts else "NO TITLE FOUND"
    except Exception as e:
        return f"ERROR: {str(e)}"

def audit_template(template, templates_dir):
    """Audit a single template"""
    file_path = templates_dir / template.file_path
    
    if not file_path.exists():
        return {
            'id': template.id,
            'name': template.name,
            'file_path': template.file_path,
            'status': 'MISSING',
            'file_title': 'FILE NOT FOUND',
            'match': False
        }
    
    # Extract title based on file extension
    if template.file_path.endswith('.docx'):
        file_title = extract_docx_title(file_path)
    elif template.file_path.endswith('.xlsx'):
        file_title = extract_xlsx_title(file_path)
    else:
        file_title = "UNKNOWN FORMAT"
    
    # Check if template name appears in file title
    template_name_clean = template.name.upper().replace('-', ' ').replace('_', ' ')
    file_title_clean = file_title.upper().replace('-', ' ').replace('_', ' ')
    
    # Simple match: check if key words from template name appear in file
    name_words = set(template_name_clean.split())
    # Remove common words
    name_words = name_words - {'THE', 'A', 'AN', 'AND', 'OR', 'FOR', 'WITH', 'PLAN', 'TEMPLATE'}
    
    match = any(word in file_title_clean for word in name_words if len(word) > 3)
    
    return {
        'id': template.id,
        'name': template.name,
        'file_path': template.file_path,
        'status': 'OK' if match else 'MISMATCH',
        'file_title': file_title,
        'match': match
    }

def main():
    templates_dir = Path(__file__).parent / 'static' / 'templates'
    
    print("=" * 80)
    print("TEMPLATE AUDIT - Checking all 955 templates")
    print("=" * 80)
    print()
    
    with app.app_context():
        templates = Template.query.all()
        
        total = len(templates)
        missing = 0
        mismatch = 0
        ok = 0
        
        mismatches = []
        
        for i, template in enumerate(templates, 1):
            result = audit_template(template, templates_dir)
            
            if result['status'] == 'MISSING':
                missing += 1
                print(f"[{i}/{total}] ❌ MISSING: ID {result['id']} - {result['name']}")
            elif result['status'] == 'MISMATCH':
                mismatch += 1
                mismatches.append(result)
                print(f"[{i}/{total}] ⚠️  MISMATCH: ID {result['id']} - {result['name']}")
                print(f"           File title: {result['file_title'][:80]}")
            else:
                ok += 1
                if i % 50 == 0:  # Print progress every 50 templates
                    print(f"[{i}/{total}] ✅ Checked {i} templates...")
        
        print()
        print("=" * 80)
        print("AUDIT SUMMARY")
        print("=" * 80)
        print(f"Total templates: {total}")
        print(f"✅ OK: {ok}")
        print(f"⚠️  MISMATCH: {mismatch}")
        print(f"❌ MISSING: {missing}")
        print()
        
        if mismatches:
            print("=" * 80)
            print("DETAILED MISMATCH REPORT")
            print("=" * 80)
            for result in mismatches[:20]:  # Show first 20 mismatches
                print(f"\nID {result['id']}: {result['name']}")
                print(f"  File: {result['file_path']}")
                print(f"  Content: {result['file_title'][:100]}")
            
            if len(mismatches) > 20:
                print(f"\n... and {len(mismatches) - 20} more mismatches")
        
        # Save full report to file
        report_path = Path(__file__).parent / 'template_audit_report.txt'
        with open(report_path, 'w') as f:
            f.write("TEMPLATE AUDIT REPORT\n")
            f.write("=" * 80 + "\n\n")
            f.write(f"Total templates: {total}\n")
            f.write(f"OK: {ok}\n")
            f.write(f"MISMATCH: {mismatch}\n")
            f.write(f"MISSING: {missing}\n\n")
            
            if mismatches:
                f.write("MISMATCHED TEMPLATES:\n")
                f.write("=" * 80 + "\n\n")
                for result in mismatches:
                    f.write(f"ID {result['id']}: {result['name']}\n")
                    f.write(f"  File: {result['file_path']}\n")
                    f.write(f"  Content: {result['file_title']}\n\n")
        
        print(f"\nFull report saved to: {report_path}")

if __name__ == '__main__':
    main()

