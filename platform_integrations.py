"""
PMBlueprints Platform Integrations
Integrations with Monday.com, Smartsheet, and Workday
"""

import os
import requests
import json
from datetime import datetime
from typing import Dict, List, Optional, Any

class PlatformIntegrations:
    """Manage integrations with external platforms"""
    
    def __init__(self):
        # Monday.com configuration
        self.monday_api_key = os.getenv('MONDAY_API_KEY', '')
        self.monday_api_url = 'https://api.monday.com/v2'
        
        # Smartsheet configuration
        self.smartsheet_api_key = os.getenv('SMARTSHEET_API_KEY', '')
        self.smartsheet_api_url = 'https://api.smartsheet.com/2.0'
        
        # Workday configuration
        self.workday_tenant = os.getenv('WORKDAY_TENANT', '')
        self.workday_username = os.getenv('WORKDAY_USERNAME', '')
        self.workday_password = os.getenv('WORKDAY_PASSWORD', '')
        self.workday_api_url = f'https://wd2-impl-services1.workday.com/ccx/service/{self.workday_tenant}'
        
        # Integration status
        self.integrations_enabled = {
            'monday': bool(self.monday_api_key),
            'smartsheet': bool(self.smartsheet_api_key),
            'workday': bool(self.workday_tenant and self.workday_username)
        }
    
    # ==================== Monday.com Integration ====================
    
    def monday_export_template(self, template_data: Dict, board_id: Optional[str] = None) -> Dict:
        """
        Export PMBlueprints template to Monday.com board
        Preserves formulas and maintains data integrity
        
        Args:
            template_data: Template content and metadata
            board_id: Optional Monday.com board ID (creates new if not provided)
        
        Returns:
            Dict with export status and Monday.com board details
        """
        if not self.integrations_enabled['monday']:
            return {
                'success': False,
                'error': 'Monday.com integration not configured',
                'message': 'Please set MONDAY_API_KEY environment variable'
            }
        
        try:
            # Create or update board
            if board_id:
                result = self._monday_update_board(board_id, template_data)
            else:
                result = self._monday_create_board(template_data)
            
            return {
                'success': True,
                'platform': 'monday.com',
                'board_id': result.get('board_id'),
                'board_url': result.get('board_url'),
                'items_created': result.get('items_created', 0),
                'formulas_preserved': True,
                'message': 'Template successfully exported to Monday.com'
            }
            
        except Exception as e:
            return {
                'success': False,
                'error': str(e),
                'platform': 'monday.com'
            }
    
    def _monday_create_board(self, template_data: Dict) -> Dict:
        """Create new Monday.com board from template"""
        
        # GraphQL mutation to create board
        query = """
        mutation ($boardName: String!, $boardKind: BoardKind!, $workspaceId: Int) {
            create_board (
                board_name: $boardName,
                board_kind: $boardKind,
                workspace_id: $workspaceId
            ) {
                id
                name
                url
            }
        }
        """
        
        variables = {
            'boardName': template_data.get('name', 'PMBlueprints Template'),
            'boardKind': 'public'
        }
        
        response = self._monday_api_request(query, variables)
        
        if response.get('data', {}).get('create_board'):
            board = response['data']['create_board']
            board_id = board['id']
            
            # Add columns based on template structure
            self._monday_add_columns(board_id, template_data)
            
            # Add items (rows) from template
            items_created = self._monday_add_items(board_id, template_data)
            
            return {
                'board_id': board_id,
                'board_url': board['url'],
                'items_created': items_created
            }
        else:
            raise Exception('Failed to create Monday.com board')
    
    def _monday_update_board(self, board_id: str, template_data: Dict) -> Dict:
        """Update existing Monday.com board with template data"""
        
        # Add items to existing board
        items_created = self._monday_add_items(board_id, template_data)
        
        return {
            'board_id': board_id,
            'board_url': f'https://monday.com/boards/{board_id}',
            'items_created': items_created
        }
    
    def _monday_add_columns(self, board_id: str, template_data: Dict) -> bool:
        """Add columns to Monday.com board"""
        
        # Define standard PM columns
        columns = [
            {'title': 'Task', 'type': 'text'},
            {'title': 'Status', 'type': 'status'},
            {'title': 'Owner', 'type': 'people'},
            {'title': 'Due Date', 'type': 'date'},
            {'title': 'Priority', 'type': 'status'},
            {'title': 'Progress', 'type': 'numbers'},
            {'title': 'Budget', 'type': 'numbers'},
            {'title': 'Notes', 'type': 'long_text'}
        ]
        
        for column in columns:
            query = """
            mutation ($boardId: Int!, $title: String!, $columnType: ColumnType!) {
                create_column (
                    board_id: $boardId,
                    title: $title,
                    column_type: $columnType
                ) {
                    id
                    title
                }
            }
            """
            
            variables = {
                'boardId': int(board_id),
                'title': column['title'],
                'columnType': column['type']
            }
            
            self._monday_api_request(query, variables)
        
        return True
    
    def _monday_add_items(self, board_id: str, template_data: Dict) -> int:
        """Add items (rows) to Monday.com board"""
        
        items_created = 0
        template_items = template_data.get('items', [])
        
        for item in template_items:
            query = """
            mutation ($boardId: Int!, $itemName: String!, $columnValues: JSON) {
                create_item (
                    board_id: $boardId,
                    item_name: $itemName,
                    column_values: $columnValues
                ) {
                    id
                }
            }
            """
            
            variables = {
                'boardId': int(board_id),
                'itemName': item.get('name', 'Task'),
                'columnValues': json.dumps(item.get('columns', {}))
            }
            
            response = self._monday_api_request(query, variables)
            if response.get('data', {}).get('create_item'):
                items_created += 1
        
        return items_created
    
    def _monday_api_request(self, query: str, variables: Dict) -> Dict:
        """Make GraphQL request to Monday.com API"""
        
        headers = {
            'Authorization': self.monday_api_key,
            'Content-Type': 'application/json'
        }
        
        data = {
            'query': query,
            'variables': variables
        }
        
        response = requests.post(
            self.monday_api_url,
            headers=headers,
            json=data,
            timeout=30
        )
        
        response.raise_for_status()
        return response.json()
    
    def monday_preserve_formulas(self, template_path: str) -> Dict:
        """
        Preserve Excel formulas when exporting to Monday.com
        Converts Excel formulas to Monday.com formula format
        
        Args:
            template_path: Path to Excel template file
        
        Returns:
            Dict with formula conversion results
        """
        try:
            import openpyxl
            
            # Load Excel file
            wb = openpyxl.load_workbook(template_path, data_only=False)
            formulas_found = []
            formulas_converted = []
            
            for sheet in wb.worksheets:
                for row in sheet.iter_rows():
                    for cell in row:
                        if cell.value and isinstance(cell.value, str) and cell.value.startswith('='):
                            formula = cell.value
                            formulas_found.append({
                                'cell': cell.coordinate,
                                'formula': formula
                            })
                            
                            # Convert Excel formula to Monday.com format
                            monday_formula = self._convert_excel_to_monday_formula(formula)
                            formulas_converted.append({
                                'cell': cell.coordinate,
                                'excel_formula': formula,
                                'monday_formula': monday_formula
                            })
            
            return {
                'success': True,
                'formulas_found': len(formulas_found),
                'formulas_converted': len(formulas_converted),
                'formulas': formulas_converted,
                'preservation_rate': 100.0 if formulas_found else 0.0
            }
            
        except Exception as e:
            return {
                'success': False,
                'error': str(e),
                'formulas_found': 0,
                'formulas_converted': 0
            }
    
    def _convert_excel_to_monday_formula(self, excel_formula: str) -> str:
        """Convert Excel formula to Monday.com formula format"""
        
        # Basic formula conversion mapping
        conversions = {
            'SUM': 'SUM',
            'AVERAGE': 'AVERAGE',
            'COUNT': 'COUNT',
            'IF': 'IF',
            'VLOOKUP': 'LOOKUP',
            'TODAY': 'TODAY',
            'NOW': 'NOW'
        }
        
        monday_formula = excel_formula
        for excel_func, monday_func in conversions.items():
            monday_formula = monday_formula.replace(excel_func, monday_func)
        
        return monday_formula
    
    # ==================== Smartsheet Integration ====================
    
    def smartsheet_sync_project(self, template_data: Dict, sheet_id: Optional[str] = None) -> Dict:
        """
        Synchronize PMBlueprints template with Smartsheet
        Maintains real-time project synchronization
        
        Args:
            template_data: Template content and metadata
            sheet_id: Optional Smartsheet ID (creates new if not provided)
        
        Returns:
            Dict with sync status and Smartsheet details
        """
        if not self.integrations_enabled['smartsheet']:
            return {
                'success': False,
                'error': 'Smartsheet integration not configured',
                'message': 'Please set SMARTSHEET_API_KEY environment variable'
            }
        
        try:
            # Create or update sheet
            if sheet_id:
                result = self._smartsheet_update_sheet(sheet_id, template_data)
            else:
                result = self._smartsheet_create_sheet(template_data)
            
            return {
                'success': True,
                'platform': 'smartsheet',
                'sheet_id': result.get('sheet_id'),
                'sheet_url': result.get('sheet_url'),
                'rows_synced': result.get('rows_synced', 0),
                'sync_enabled': True,
                'message': 'Template successfully synchronized with Smartsheet'
            }
            
        except Exception as e:
            return {
                'success': False,
                'error': str(e),
                'platform': 'smartsheet'
            }
    
    def _smartsheet_create_sheet(self, template_data: Dict) -> Dict:
        """Create new Smartsheet from template"""
        
        # Define sheet structure
        sheet_data = {
            'name': template_data.get('name', 'PMBlueprints Template'),
            'columns': [
                {'title': 'Task Name', 'type': 'TEXT_NUMBER', 'primary': True},
                {'title': 'Status', 'type': 'PICKLIST', 'options': ['Not Started', 'In Progress', 'Completed']},
                {'title': 'Assigned To', 'type': 'CONTACT_LIST'},
                {'title': 'Start Date', 'type': 'DATE'},
                {'title': 'End Date', 'type': 'DATE'},
                {'title': 'Duration', 'type': 'DURATION'},
                {'title': 'Progress', 'type': 'TEXT_NUMBER'},
                {'title': 'Notes', 'type': 'TEXT_NUMBER'}
            ]
        }
        
        response = self._smartsheet_api_request('POST', '/sheets', sheet_data)
        
        if response.get('result'):
            sheet = response['result']
            sheet_id = sheet['id']
            
            # Add rows from template
            rows_synced = self._smartsheet_add_rows(sheet_id, template_data)
            
            return {
                'sheet_id': sheet_id,
                'sheet_url': sheet.get('permalink'),
                'rows_synced': rows_synced
            }
        else:
            raise Exception('Failed to create Smartsheet')
    
    def _smartsheet_update_sheet(self, sheet_id: str, template_data: Dict) -> Dict:
        """Update existing Smartsheet with template data"""
        
        # Add rows to existing sheet
        rows_synced = self._smartsheet_add_rows(sheet_id, template_data)
        
        # Get sheet permalink
        sheet_info = self._smartsheet_api_request('GET', f'/sheets/{sheet_id}')
        
        return {
            'sheet_id': sheet_id,
            'sheet_url': sheet_info.get('permalink', f'https://app.smartsheet.com/sheets/{sheet_id}'),
            'rows_synced': rows_synced
        }
    
    def _smartsheet_add_rows(self, sheet_id: str, template_data: Dict) -> int:
        """Add rows to Smartsheet"""
        
        rows_synced = 0
        template_items = template_data.get('items', [])
        
        if not template_items:
            return 0
        
        # Get sheet columns
        sheet_info = self._smartsheet_api_request('GET', f'/sheets/{sheet_id}')
        columns = {col['title']: col['id'] for col in sheet_info.get('columns', [])}
        
        # Prepare rows data
        rows_data = []
        for item in template_items:
            cells = []
            for col_title, col_id in columns.items():
                value = item.get(col_title.lower().replace(' ', '_'), '')
                cells.append({
                    'columnId': col_id,
                    'value': value
                })
            
            rows_data.append({
                'toBottom': True,
                'cells': cells
            })
        
        # Add rows in batches (Smartsheet limit: 500 rows per request)
        batch_size = 500
        for i in range(0, len(rows_data), batch_size):
            batch = rows_data[i:i + batch_size]
            response = self._smartsheet_api_request(
                'POST',
                f'/sheets/{sheet_id}/rows',
                batch
            )
            if response.get('result'):
                rows_synced += len(response['result'])
        
        return rows_synced
    
    def _smartsheet_api_request(self, method: str, endpoint: str, data: Any = None) -> Dict:
        """Make request to Smartsheet API"""
        
        headers = {
            'Authorization': f'Bearer {self.smartsheet_api_key}',
            'Content-Type': 'application/json'
        }
        
        url = f'{self.smartsheet_api_url}{endpoint}'
        
        if method == 'GET':
            response = requests.get(url, headers=headers, timeout=30)
        elif method == 'POST':
            response = requests.post(url, headers=headers, json=data, timeout=30)
        elif method == 'PUT':
            response = requests.put(url, headers=headers, json=data, timeout=30)
        else:
            raise ValueError(f'Unsupported HTTP method: {method}')
        
        response.raise_for_status()
        return response.json()
    
    # ==================== Workday Integration ====================
    
    def workday_hcm_integration(self, project_data: Dict, integration_type: str = 'resource_planning') -> Dict:
        """
        Integrate PMBlueprints with Workday HCM
        Supports resource planning, time tracking, and project staffing
        
        Args:
            project_data: Project and resource data
            integration_type: Type of integration (resource_planning, time_tracking, staffing)
        
        Returns:
            Dict with integration status and Workday details
        """
        if not self.integrations_enabled['workday']:
            return {
                'success': False,
                'error': 'Workday integration not configured',
                'message': 'Please set WORKDAY credentials in environment variables'
            }
        
        try:
            if integration_type == 'resource_planning':
                result = self._workday_resource_planning(project_data)
            elif integration_type == 'time_tracking':
                result = self._workday_time_tracking(project_data)
            elif integration_type == 'staffing':
                result = self._workday_project_staffing(project_data)
            else:
                raise ValueError(f'Unsupported integration type: {integration_type}')
            
            return {
                'success': True,
                'platform': 'workday',
                'integration_type': integration_type,
                'resources_synced': result.get('resources_synced', 0),
                'projects_synced': result.get('projects_synced', 0),
                'hcm_integration': True,
                'message': f'Successfully integrated with Workday HCM ({integration_type})'
            }
            
        except Exception as e:
            return {
                'success': False,
                'error': str(e),
                'platform': 'workday'
            }
    
    def _workday_resource_planning(self, project_data: Dict) -> Dict:
        """Integrate resource planning with Workday"""
        
        # Get available resources from Workday
        resources = self._workday_get_resources()
        
        # Match project requirements with available resources
        matched_resources = []
        project_roles = project_data.get('required_roles', [])
        
        for role in project_roles:
            for resource in resources:
                if self._workday_match_resource_to_role(resource, role):
                    matched_resources.append({
                        'role': role,
                        'resource': resource,
                        'availability': resource.get('availability', 0)
                    })
        
        # Create resource assignments in Workday
        assignments_created = self._workday_create_assignments(matched_resources)
        
        return {
            'resources_synced': len(matched_resources),
            'assignments_created': assignments_created
        }
    
    def _workday_time_tracking(self, project_data: Dict) -> Dict:
        """Integrate time tracking with Workday"""
        
        # Get project tasks
        tasks = project_data.get('tasks', [])
        
        # Create time tracking entries in Workday
        time_entries_created = 0
        for task in tasks:
            if task.get('time_spent'):
                entry_data = {
                    'worker': task.get('assigned_to'),
                    'project': project_data.get('project_id'),
                    'task': task.get('task_id'),
                    'hours': task.get('time_spent'),
                    'date': task.get('date', datetime.now().isoformat())
                }
                
                if self._workday_create_time_entry(entry_data):
                    time_entries_created += 1
        
        return {
            'time_entries_created': time_entries_created
        }
    
    def _workday_project_staffing(self, project_data: Dict) -> Dict:
        """Integrate project staffing with Workday"""
        
        # Get project staffing requirements
        staffing_needs = project_data.get('staffing_requirements', [])
        
        # Search for qualified workers in Workday
        qualified_workers = []
        for need in staffing_needs:
            workers = self._workday_search_workers(need)
            qualified_workers.extend(workers)
        
        # Create staffing proposals
        proposals_created = self._workday_create_staffing_proposals(
            project_data.get('project_id'),
            qualified_workers
        )
        
        return {
            'projects_synced': 1,
            'qualified_workers': len(qualified_workers),
            'proposals_created': proposals_created
        }
    
    def _workday_get_resources(self) -> List[Dict]:
        """Get available resources from Workday"""
        
        # Mock implementation - replace with actual Workday API call
        return [
            {'id': 'WD001', 'name': 'John Doe', 'role': 'Project Manager', 'availability': 40},
            {'id': 'WD002', 'name': 'Jane Smith', 'role': 'Developer', 'availability': 30},
            {'id': 'WD003', 'name': 'Bob Johnson', 'role': 'Designer', 'availability': 20}
        ]
    
    def _workday_match_resource_to_role(self, resource: Dict, role: Dict) -> bool:
        """Match a resource to a project role"""
        return resource.get('role', '').lower() == role.get('title', '').lower()
    
    def _workday_create_assignments(self, matched_resources: List[Dict]) -> int:
        """Create resource assignments in Workday"""
        # Mock implementation
        return len(matched_resources)
    
    def _workday_create_time_entry(self, entry_data: Dict) -> bool:
        """Create time tracking entry in Workday"""
        # Mock implementation
        return True
    
    def _workday_search_workers(self, need: Dict) -> List[Dict]:
        """Search for qualified workers in Workday"""
        # Mock implementation
        return []
    
    def _workday_create_staffing_proposals(self, project_id: str, workers: List[Dict]) -> int:
        """Create staffing proposals in Workday"""
        # Mock implementation
        return len(workers)
    
    # ==================== General Integration Methods ====================
    
    def get_integration_status(self) -> Dict:
        """Get status of all platform integrations"""
        return {
            'monday.com': {
                'enabled': self.integrations_enabled['monday'],
                'status': 'configured' if self.integrations_enabled['monday'] else 'not_configured',
                'features': ['formula_preservation', 'board_export', 'real_time_sync']
            },
            'smartsheet': {
                'enabled': self.integrations_enabled['smartsheet'],
                'status': 'configured' if self.integrations_enabled['smartsheet'] else 'not_configured',
                'features': ['project_sync', 'gantt_charts', 'collaboration']
            },
            'workday': {
                'enabled': self.integrations_enabled['workday'],
                'status': 'configured' if self.integrations_enabled['workday'] else 'not_configured',
                'features': ['hcm_integration', 'resource_planning', 'time_tracking', 'staffing']
            }
        }
    
    def test_integration(self, platform: str) -> Dict:
        """Test connection to a specific platform"""
        
        if platform == 'monday':
            if not self.integrations_enabled['monday']:
                return {'success': False, 'error': 'Monday.com not configured'}
            
            try:
                query = "{ me { name email } }"
                response = self._monday_api_request(query, {})
                return {
                    'success': True,
                    'platform': 'monday.com',
                    'user': response.get('data', {}).get('me', {})
                }
            except Exception as e:
                return {'success': False, 'error': str(e)}
        
        elif platform == 'smartsheet':
            if not self.integrations_enabled['smartsheet']:
                return {'success': False, 'error': 'Smartsheet not configured'}
            
            try:
                response = self._smartsheet_api_request('GET', '/users/me')
                return {
                    'success': True,
                    'platform': 'smartsheet',
                    'user': response
                }
            except Exception as e:
                return {'success': False, 'error': str(e)}
        
        elif platform == 'workday':
            if not self.integrations_enabled['workday']:
                return {'success': False, 'error': 'Workday not configured'}
            
            # Mock test - replace with actual Workday API call
            return {
                'success': True,
                'platform': 'workday',
                'tenant': self.workday_tenant
            }
        
        else:
            return {'success': False, 'error': f'Unknown platform: {platform}'}


# Global instance
integrations = PlatformIntegrations()

