#!/usr/bin/env python3
"""
Generate proper Excel screenshot from sheet 2 (dashboard) in landscape orientation
showing ALL columns visible.
"""

import openpyxl
from openpyxl.utils import get_column_letter
from PIL import Image, ImageDraw, ImageFont
import os
import sys

def generate_excel_screenshot(excel_path, output_path, sheet_index=1):
    """
    Generate a landscape screenshot of an Excel file's specified sheet.
    
    Args:
        excel_path: Path to the Excel file
        output_path: Path to save the screenshot
        sheet_index: Index of sheet to capture (0-based, default 1 for second sheet)
    """
    print(f"Loading Excel file: {excel_path}")
    
    # Load the workbook
    wb = openpyxl.load_workbook(excel_path, data_only=True)
    
    # Get the sheet names
    sheet_names = wb.sheetnames
    print(f"Available sheets: {sheet_names}")
    
    if sheet_index >= len(sheet_names):
        print(f"Warning: Sheet index {sheet_index} not found, using first sheet")
        sheet_index = 0
    
    sheet = wb.worksheets[sheet_index]
    print(f"Using sheet: {sheet.title} (index {sheet_index})")
    
    # Get the used range
    max_row = sheet.max_row
    max_col = sheet.max_column
    print(f"Sheet dimensions: {max_row} rows x {max_col} columns")
    
    # Calculate image dimensions (landscape)
    # Use larger dimensions for better quality
    cell_width = 120  # pixels per cell
    cell_height = 30  # pixels per cell
    
    # Limit to reasonable size (first 20 rows, all columns)
    display_rows = min(max_row, 20)
    display_cols = max_col
    
    img_width = display_cols * cell_width
    img_height = display_rows * cell_height + 100  # Extra space for header
    
    print(f"Creating image: {img_width}x{img_height} pixels (landscape)")
    
    # Create image with white background
    img = Image.new('RGB', (img_width, img_height), color='white')
    draw = ImageDraw.Draw(img)
    
    # Try to use a nice font, fallback to default
    try:
        font = ImageFont.truetype("/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf", 12)
        header_font = ImageFont.truetype("/usr/share/fonts/truetype/dejavu/DejaVuSans-Bold.ttf", 14)
    except:
        font = ImageFont.load_default()
        header_font = ImageFont.load_default()
    
    # Draw sheet name header
    header_text = f"Sheet: {sheet.title}"
    draw.rectangle([0, 0, img_width, 40], fill='#2c3e50')
    draw.text((10, 10), header_text, fill='white', font=header_font)
    
    y_offset = 50
    
    # Draw grid and cell contents
    for row_idx in range(1, display_rows + 1):
        for col_idx in range(1, display_cols + 1):
            cell = sheet.cell(row=row_idx, column=col_idx)
            
            # Calculate cell position
            x = (col_idx - 1) * cell_width
            y = y_offset + (row_idx - 1) * cell_height
            
            # Draw cell border
            draw.rectangle([x, y, x + cell_width, y + cell_height], 
                          outline='#cccccc', width=1)
            
            # Get cell value
            value = cell.value
            if value is not None:
                # Convert to string and truncate if too long
                text = str(value)
                if len(text) > 15:
                    text = text[:12] + "..."
                
                # Determine background color based on fill
                bg_color = 'white'
                if cell.fill and cell.fill.start_color:
                    try:
                        color_hex = cell.fill.start_color.rgb
                        if color_hex and len(color_hex) == 8:  # ARGB format
                            bg_color = '#' + color_hex[2:]  # Skip alpha
                            # Fill cell background
                            draw.rectangle([x+1, y+1, x + cell_width-1, y + cell_height-1], 
                                         fill=bg_color)
                    except:
                        pass
                
                # Draw text
                text_color = 'black'
                draw.text((x + 5, y + 8), text, fill=text_color, font=font)
    
    # Save the image
    img.save(output_path, 'PNG', quality=95)
    print(f"Screenshot saved to: {output_path}")
    print(f"Image size: {img.width}x{img.height} pixels")
    
    return output_path


if __name__ == "__main__":
    # Test with KPI Dashboard file
    test_file = "/home/ubuntu/pmb_repo/static/templates/AI_ML_KPI_Dashboard_with_Instructions_2025_PMI.xlsx"
    output_file = "/home/ubuntu/pmb_repo/sample_screenshot_sheet2.png"
    
    if len(sys.argv) > 1:
        test_file = sys.argv[1]
    if len(sys.argv) > 2:
        output_file = sys.argv[2]
    
    if not os.path.exists(test_file):
        print(f"Error: File not found: {test_file}")
        sys.exit(1)
    
    generate_excel_screenshot(test_file, output_file, sheet_index=1)
    print("\n✓ Sample screenshot generated successfully!")
    print(f"  View at: {output_file}")

