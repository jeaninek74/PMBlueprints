"""
Thumbnail Generator
Generates real thumbnail images from actual template files (Excel, Word, PowerPoint)
"""

import os
import logging
from PIL import Image, ImageDraw, ImageFont
from pdf2image import convert_from_path
from docx import Document
from openpyxl import load_workbook
from openpyxl.drawing.image import Image as XLImage
import tempfile
import subprocess

logger = logging.getLogger(__name__)

class ThumbnailGenerator:
    """Generate thumbnails from template files"""
    
    def __init__(self, thumbnail_dir='static/thumbnails'):
        self.thumbnail_dir = thumbnail_dir
        os.makedirs(thumbnail_dir, exist_ok=True)
    
    def generate_thumbnail(self, file_path, template_id, file_format):
        """
        Generate thumbnail from template file
        Returns: path to generated thumbnail or None if failed
        """
        try:
            if file_format.lower() in ['xlsx', 'xls']:
                return self._generate_excel_thumbnail(file_path, template_id)
            elif file_format.lower() in ['docx', 'doc']:
                return self._generate_word_thumbnail(file_path, template_id)
            elif file_format.lower() in ['pptx', 'ppt']:
                return self._generate_powerpoint_thumbnail(file_path, template_id)
            else:
                logger.warning(f"Unsupported format: {file_format}")
                return None
        except Exception as e:
            logger.error(f"Error generating thumbnail for {file_path}: {str(e)}")
            return None
    
    def _generate_excel_thumbnail(self, file_path, template_id):
        """Generate thumbnail from Excel file"""
        try:
            # Load workbook
            wb = load_workbook(file_path, data_only=True)
            ws = wb.active
            
            # Create image with white background
            img_width, img_height = 400, 300
            img = Image.new('RGB', (img_width, img_height), 'white')
            draw = ImageDraw.Draw(img)
            
            # Try to load a font
            try:
                font = ImageFont.truetype("/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf", 12)
                title_font = ImageFont.truetype("/usr/share/fonts/truetype/dejavu/DejaVuSans-Bold.ttf", 16)
            except:
                font = ImageFont.load_default()
                title_font = ImageFont.load_default()
            
            # Draw Excel-like grid and content
            y_offset = 20
            
            # Draw title (sheet name)
            draw.text((10, y_offset), ws.title, fill='black', font=title_font)
            y_offset += 40
            
            # Draw grid with cell values
            row_height = 25
            col_width = 100
            
            for row_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=10, max_col=4, values_only=True)):
                if y_offset > img_height - 30:
                    break
                    
                x_offset = 10
                for col_idx, cell_value in enumerate(row):
                    # Draw cell border
                    draw.rectangle(
                        [x_offset, y_offset, x_offset + col_width, y_offset + row_height],
                        outline='gray'
                    )
                    
                    # Draw cell value
                    if cell_value:
                        text = str(cell_value)[:15]  # Truncate long text
                        draw.text((x_offset + 5, y_offset + 5), text, fill='black', font=font)
                    
                    x_offset += col_width
                
                y_offset += row_height
            
            # Add Excel logo/indicator
            draw.rectangle([img_width - 60, 10, img_width - 10, 35], fill='#217346')
            draw.text((img_width - 55, 15), 'XLSX', fill='white', font=font)
            
            # Save thumbnail
            thumbnail_path = os.path.join(self.thumbnail_dir, f'template_{template_id}.png')
            img.save(thumbnail_path, 'PNG')
            
            return thumbnail_path
            
        except Exception as e:
            logger.error(f"Error generating Excel thumbnail: {str(e)}")
            return None
    
    def _generate_word_thumbnail(self, file_path, template_id):
        """Generate thumbnail from Word document"""
        try:
            # Try to convert Word to PDF first, then to image
            # This requires LibreOffice or similar
            with tempfile.TemporaryDirectory() as temp_dir:
                # Convert DOCX to PDF using LibreOffice
                pdf_path = os.path.join(temp_dir, 'temp.pdf')
                
                try:
                    subprocess.run([
                        'libreoffice',
                        '--headless',
                        '--convert-to', 'pdf',
                        '--outdir', temp_dir,
                        file_path
                    ], check=True, timeout=30, capture_output=True)
                    
                    # Find the generated PDF
                    pdf_files = [f for f in os.listdir(temp_dir) if f.endswith('.pdf')]
                    if pdf_files:
                        pdf_path = os.path.join(temp_dir, pdf_files[0])
                        
                        # Convert first page of PDF to image
                        images = convert_from_path(pdf_path, first_page=1, last_page=1, dpi=150)
                        if images:
                            # Resize to thumbnail size
                            img = images[0]
                            img.thumbnail((400, 300), Image.Resampling.LANCZOS)
                            
                            # Create white background
                            thumbnail = Image.new('RGB', (400, 300), 'white')
                            # Center the image
                            offset = ((400 - img.width) // 2, (300 - img.height) // 2)
                            thumbnail.paste(img, offset)
                            
                            # Save thumbnail
                            thumbnail_path = os.path.join(self.thumbnail_dir, f'template_{template_id}.png')
                            thumbnail.save(thumbnail_path, 'PNG')
                            
                            return thumbnail_path
                
                except subprocess.TimeoutExpired:
                    logger.warning("LibreOffice conversion timeout")
                except subprocess.CalledProcessError as e:
                    logger.warning(f"LibreOffice conversion failed: {e}")
            
            # Fallback: Read document and create text-based thumbnail
            doc = Document(file_path)
            
            # Create image
            img = Image.new('RGB', (400, 300), 'white')
            draw = ImageDraw.Draw(img)
            
            try:
                font = ImageFont.truetype("/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf", 11)
                title_font = ImageFont.truetype("/usr/share/fonts/truetype/dejavu/DejaVuSans-Bold.ttf", 14)
            except:
                font = ImageFont.load_default()
                title_font = ImageFont.load_default()
            
            # Draw document icon
            draw.rectangle([10, 10, 50, 60], outline='blue', width=2)
            draw.polygon([(50, 10), (50, 30), (30, 30)], fill='blue')
            
            # Draw first few paragraphs
            y_offset = 80
            for para in doc.paragraphs[:8]:
                if y_offset > 280:
                    break
                text = para.text.strip()
                if text:
                    # Truncate long lines
                    if len(text) > 50:
                        text = text[:50] + '...'
                    draw.text((10, y_offset), text, fill='black', font=font)
                    y_offset += 25
            
            # Add Word logo/indicator
            draw.rectangle([350, 10, 390, 35], fill='#2B579A')
            draw.text((355, 15), 'DOCX', fill='white', font=font)
            
            # Save thumbnail
            thumbnail_path = os.path.join(self.thumbnail_dir, f'template_{template_id}.png')
            img.save(thumbnail_path, 'PNG')
            
            return thumbnail_path
            
        except Exception as e:
            logger.error(f"Error generating Word thumbnail: {str(e)}")
            return None
    
    def _generate_powerpoint_thumbnail(self, file_path, template_id):
        """Generate thumbnail from PowerPoint file"""
        try:
            # Similar to Word, try to convert to PDF then image
            with tempfile.TemporaryDirectory() as temp_dir:
                try:
                    subprocess.run([
                        'libreoffice',
                        '--headless',
                        '--convert-to', 'pdf',
                        '--outdir', temp_dir,
                        file_path
                    ], check=True, timeout=30, capture_output=True)
                    
                    # Find the generated PDF
                    pdf_files = [f for f in os.listdir(temp_dir) if f.endswith('.pdf')]
                    if pdf_files:
                        pdf_path = os.path.join(temp_dir, pdf_files[0])
                        
                        # Convert first page of PDF to image
                        images = convert_from_path(pdf_path, first_page=1, last_page=1, dpi=150)
                        if images:
                            # Resize to thumbnail size
                            img = images[0]
                            img.thumbnail((400, 300), Image.Resampling.LANCZOS)
                            
                            # Create white background
                            thumbnail = Image.new('RGB', (400, 300), 'white')
                            offset = ((400 - img.width) // 2, (300 - img.height) // 2)
                            thumbnail.paste(img, offset)
                            
                            # Save thumbnail
                            thumbnail_path = os.path.join(self.thumbnail_dir, f'template_{template_id}.png')
                            thumbnail.save(thumbnail_path, 'PNG')
                            
                            return thumbnail_path
                
                except (subprocess.TimeoutExpired, subprocess.CalledProcessError):
                    pass
            
            # Fallback: Create PowerPoint-style thumbnail
            img = Image.new('RGB', (400, 300), 'white')
            draw = ImageDraw.Draw(img)
            
            try:
                font = ImageFont.truetype("/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf", 12)
                title_font = ImageFont.truetype("/usr/share/fonts/truetype/dejavu/DejaVuSans-Bold.ttf", 16)
            except:
                font = ImageFont.load_default()
                title_font = ImageFont.load_default()
            
            # Draw slide frame
            draw.rectangle([20, 40, 380, 260], outline='gray', width=2)
            
            # Draw title placeholder
            draw.rectangle([40, 60, 360, 100], fill='#D35400', outline='#D35400')
            draw.text((50, 70), 'Presentation Template', fill='white', font=title_font)
            
            # Draw content placeholders
            draw.rectangle([40, 120, 360, 180], outline='gray')
            draw.rectangle([40, 190, 360, 240], outline='gray')
            
            # Add PowerPoint logo/indicator
            draw.rectangle([350, 10, 390, 35], fill='#D35400')
            draw.text((355, 15), 'PPTX', fill='white', font=font)
            
            # Save thumbnail
            thumbnail_path = os.path.join(self.thumbnail_dir, f'template_{template_id}.png')
            img.save(thumbnail_path, 'PNG')
            
            return thumbnail_path
            
        except Exception as e:
            logger.error(f"Error generating PowerPoint thumbnail: {str(e)}")
            return None
    
    def generate_all_thumbnails(self, templates):
        """Generate thumbnails for all templates"""
        success_count = 0
        fail_count = 0
        
        for template in templates:
            if not template.file_path or not os.path.exists(template.file_path):
                logger.warning(f"Template file not found: {template.file_path}")
                fail_count += 1
                continue
            
            thumbnail_path = self.generate_thumbnail(
                template.file_path,
                template.id,
                template.file_format
            )
            
            if thumbnail_path:
                # Update template with thumbnail path
                template.thumbnail_path = thumbnail_path
                success_count += 1
            else:
                fail_count += 1
        
        logger.info(f"Thumbnail generation complete: {success_count} success, {fail_count} failed")
        return success_count, fail_count

